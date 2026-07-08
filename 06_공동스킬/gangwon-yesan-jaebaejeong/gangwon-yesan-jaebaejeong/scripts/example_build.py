# -*- coding: utf-8 -*-
"""
예시 빌드: reallocation_lib를 이용해 6시트 재배정 워크북을 조립하는 참조 구현.

⚠️ 여기 DEMO 데이터는 스킬 동작 검증용 '플레이스홀더'다. 실제 작업에서는
   반드시 사용자가 제공한 자료로 교체한다(임의 데이터 생성 금지).

사용:  python scripts/example_build.py [출력파일.xlsx]
검증:  python scripts/recalc.py 출력파일.xlsx
"""
import sys
from openpyxl import Workbook
import reallocation_lib as L

DETAIL = "세부 재배정 내역"

# --- DEMO 데이터(플레이스홀더) : 실제 작업 시 사용자 자료로 교체 ---
DEMO = [
    {"지역청": "춘천교육지원청", "학교급": "초등학교", "설립유형": "공립",
     "학교명": "가나초등학교", "분야": "합창", "팀명": "", "지원금액": 1500},
    {"지역청": "춘천시", "학교급": "고", "설립유형": "사립",
     "학교명": "다라고등학교", "분야": "합창", "팀명": "", "지원금액": 2500},
    {"지역청": "원주교육지원청", "학교급": "중", "설립유형": "공립",
     "학교명": "마바중학교", "분야": "오케스트라", "팀명": "", "지원금액": 3000},
    {"지역청": "속양", "학교급": "초", "설립유형": "공립",
     "학교명": "사아초등학교", "분야": "아카펠라", "팀명": "", "지원금액": 1500},
]


def normalize(rows):
    out = []
    for r in rows:
        out.append({
            "지역청": L.normalize_region(r["지역청"]),
            "학교급": L.normalize_level(r["학교급"]),
            "설립유형": L.normalize_estab(r["설립유형"]),
            "학교명": r["학교명"],
            "분야": r.get("분야", ""),
            "팀명": r.get("팀명", ""),
            "지원금액": r["지원금액"],
        })
    return out


def build(rows, path):
    rows = L.sort_detail(normalize(rows))
    H = L.excel_helpers()
    wb = Workbook()

    # 1) 세부 재배정 내역  (A연번 B지역청 C학교급 D설립유형 E학교명 F분야 G팀명 H지원금액 I비고)
    ws = wb.active
    ws.title = DETAIL
    cols = ["연번", "지역청", "학교급", "설립유형", "학교명",
            "분야/부문", "팀명", "지원금액(원)", "비고"]
    H["title"](ws, "2026 예시사업 재배정 내역 (단위: 원)", len(cols))
    H["header"](ws, 2, cols)
    first = 3
    for i, r in enumerate(rows):
        rr = first + i
        vals = [i + 1, r["지역청"], r["학교급"], r["설립유형"], r["학교명"],
                r["분야"], r["팀명"], r["지원금액"], ""]
        for j, v in enumerate(vals, 1):
            H["cell"](ws, rr, j, v, num=(j == 8))
    last = first + len(rows) - 1
    tot = last + 1
    H["cell"](ws, tot, 1, "합계", bold=True, total=True)
    for j in range(2, 8):
        H["cell"](ws, tot, j, "", total=True)
    H["cell"](ws, tot, 8, f"=SUM(H{first}:H{last})", num=True, total=True)
    H["cell"](ws, tot, 9, "", total=True)
    H["autofit"](ws, [6, 12, 8, 10, 20, 12, 12, 14, 16])
    H["landscape_fit"](ws)
    dref = f"'{DETAIL}'"

    # 2) 지역청별 재배정 내역
    regions = L.order_regions([r["지역청"] for r in rows])
    ws2 = wb.create_sheet("지역청별 재배정 내역")
    c2 = ["지역청", "초", "중", "고", "특수/기타", "합계", "비고"]
    H["title"](ws2, "지역청별 재배정 내역 (단위: 원)", len(c2))
    H["header"](ws2, 2, c2)
    for k, reg in enumerate(regions):
        r0 = 3 + k
        H["cell"](ws2, r0, 1, reg)
        for col, lv in zip(range(2, 6), ["초", "중", "고", "특수"]):
            f = (f"=SUMIFS({dref}!H:H,{dref}!B:B,A{r0},{dref}!C:C,\"{lv}\")")
            H["cell"](ws2, r0, col, f, num=True)
        H["cell"](ws2, r0, 6, f"=SUM(B{r0}:E{r0})", num=True)
        H["cell"](ws2, r0, 7, "")
    rt = 3 + len(regions)
    H["cell"](ws2, rt, 1, "합계", bold=True, total=True)
    for col in range(2, 7):
        cl = chr(64 + col)
        H["cell"](ws2, rt, col, f"=SUM({cl}3:{cl}{rt-1})", num=True, total=True)
    H["cell"](ws2, rt, 7, "", total=True)
    H["autofit"](ws2, [12, 12, 12, 12, 12, 14, 16])

    # 3) 학교급별 재배정 내역
    ws3 = wb.create_sheet("학교급별 재배정 내역")
    c3 = ["학교급", "국립", "공립", "사립", "기타", "전체", "지원금액 합계", "비고"]
    H["title"](ws3, "학교급별 재배정 내역 (단위: 원)", len(c3))
    H["header"](ws3, 2, c3)
    for k, lv in enumerate(L.LEVEL_ORDER):
        r0 = 3 + k
        H["cell"](ws3, r0, 1, lv)
        for col, es in zip(range(2, 6), L.ESTAB_ORDER):
            f = (f"=COUNTIFS({dref}!C:C,A{r0},{dref}!D:D,\"{es}\")")
            H["cell"](ws3, r0, col, f, num=True)
        H["cell"](ws3, r0, 6, f"=SUM(B{r0}:E{r0})", num=True)
        H["cell"](ws3, r0, 7, f"=SUMIFS({dref}!H:H,{dref}!C:C,A{r0})", num=True)
        H["cell"](ws3, r0, 8, "")
    H["autofit"](ws3, [10, 8, 8, 8, 8, 8, 16, 16])

    # 4) 설립유형별 재배정 내역
    ws4 = wb.create_sheet("설립유형별 재배정 내역")
    c4 = ["설립유형", "초", "중", "고", "특수/기타", "지원금액 합계", "비고"]
    H["title"](ws4, "설립유형별 재배정 내역 (단위: 원)", len(c4))
    H["header"](ws4, 2, c4)
    for k, es in enumerate(L.ESTAB_ORDER):
        r0 = 3 + k
        H["cell"](ws4, r0, 1, es)
        for col, lv in zip(range(2, 6), ["초", "중", "고", "특수"]):
            f = (f"=COUNTIFS({dref}!D:D,A{r0},{dref}!C:C,\"{lv}\")")
            H["cell"](ws4, r0, col, f, num=True)
        H["cell"](ws4, r0, 6, f"=SUMIFS({dref}!H:H,{dref}!D:D,A{r0})", num=True)
        H["cell"](ws4, r0, 7, "")
    H["autofit"](ws4, [12, 8, 8, 8, 12, 16, 16])

    # 5) 재배정 총괄
    ws0 = wb.create_sheet("재배정 총괄", 0)
    H["title"](ws0, "2026 예시사업 예산 재배정 총괄", 2)
    items = [
        ("사업명", "예시사업(플레이스홀더)"),
        ("연도", "2026"),
        ("재배정 총액(원)", f"='{DETAIL}'!H{tot}"),
        ("지역청 수", len(regions)),
        ("지원 학교 수", len(rows)),
    ]
    for k, (label, val) in enumerate(items):
        r0 = 3 + k
        H["cell"](ws0, r0, 1, label, bold=True)
        H["cell"](ws0, r0, 2, val, num=(label == "재배정 총액(원)"))
    H["autofit"](ws0, [18, 30])

    # 6) 검토표
    ws5 = wb.create_sheet("검토표")
    c5 = ["구분", "검토 항목", "결과", "확인 내용", "조치 필요"]
    H["title"](ws5, "검토표", len(c5))
    H["header"](ws5, 2, c5)
    detail_total = sum(r["지원금액"] for r in rows)
    checklist = L.build_full_checklist(
        rows, allowed_regions=regions, detail_total=detail_total,
        region_total=detail_total, level_total=detail_total,
        estab_total=detail_total, plan_total=detail_total)
    for k, row in enumerate(checklist):
        r0 = 3 + k
        for j, v in enumerate(row, 1):
            H["cell"](ws5, r0, j, v)
    H["autofit"](ws5, [10, 34, 12, 40, 10])

    wb.save(path)
    return path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "재배정_예시_정리본.xlsx"
    print("생성:", build(DEMO, out))
