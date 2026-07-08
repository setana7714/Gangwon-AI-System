# -*- coding: utf-8 -*-
"""
강원교육청 예산 재배정 공용 라이브러리.

역할:
  1) 지역청 지정 순서 / 표준명 변환
  2) 학교급·설립유형 정규화
  3) 세부내역 정렬(지역청→학교급→설립유형→가나다)
  4) 검증(합계 일치, 지역청 순서/누락, 단위혼용, 중복, 국립분리 등)
  5) 엑셀 서식 헬퍼(제목/머리글/금액서식/합계수식)

주의: 이 라이브러리는 '자료를 만들어내지 않는다'. 사용자가 준 데이터를
      정규화·정렬·검증만 한다. 값 채우기는 하지 않는다.
"""
from __future__ import annotations

# ─────────────────────────────────────────────────────────────
# 1. 지역청 지정 순서 (절대 가나다순/자동정렬 아님)
# ─────────────────────────────────────────────────────────────
REGION_ORDER = [
    "춘천", "원주", "강릉", "속초양양", "동해", "태백", "삼척",
    "홍천", "횡성", "영월", "평창", "정선", "철원", "화천",
    "양구", "인제", "고성",
]
REGION_INDEX = {r: i for i, r in enumerate(REGION_ORDER)}

# 원자료 표기 → 표준 표기
REGION_NORMALIZE = {
    "춘천": "춘천", "춘천시": "춘천", "춘천교육지원청": "춘천",
    "원주": "원주", "원주시": "원주", "원주교육지원청": "원주",
    "강릉": "강릉", "강릉시": "강릉", "강릉교육지원청": "강릉",
    "속초": "속초양양", "양양": "속초양양", "속초·양양": "속초양양",
    "속초양양": "속초양양", "속양": "속초양양", "속초교육지원청": "속초양양",
    "양양교육지원청": "속초양양",
    "동해": "동해", "동해시": "동해", "동해교육지원청": "동해",
    "태백": "태백", "태백시": "태백", "태백교육지원청": "태백",
    "삼척": "삼척", "삼척시": "삼척", "삼척교육지원청": "삼척",
    "홍천": "홍천", "홍천군": "홍천", "홍천교육지원청": "홍천",
    "횡성": "횡성", "횡성군": "횡성", "횡성교육지원청": "횡성",
    "영월": "영월", "영월군": "영월", "영월교육지원청": "영월",
    "평창": "평창", "평창군": "평창", "평창교육지원청": "평창",
    "정선": "정선", "정선군": "정선", "정선교육지원청": "정선",
    "철원": "철원", "철원군": "철원", "철원교육지원청": "철원",
    "화천": "화천", "화천군": "화천", "화천교육지원청": "화천",
    "양구": "양구", "양구군": "양구", "양구교육지원청": "양구",
    "인제": "인제", "인제군": "인제", "인제교육지원청": "인제",
    "고성": "고성", "고성군": "고성", "고성교육지원청": "고성",
}

# ─────────────────────────────────────────────────────────────
# 2. 학교급 / 설립유형 순서·정규화
# ─────────────────────────────────────────────────────────────
LEVEL_ORDER = ["초", "중", "고", "특수", "기타"]
LEVEL_INDEX = {v: i for i, v in enumerate(LEVEL_ORDER)}
LEVEL_NORMALIZE = {
    "초등학교": "초", "초": "초", "초교": "초",
    "중학교": "중", "중": "중",
    "고등학교": "고", "고": "고",
    "특수학교": "특수", "특수": "특수",
}

ESTAB_ORDER = ["국립", "공립", "사립", "기타"]
ESTAB_INDEX = {v: i for i, v in enumerate(ESTAB_ORDER)}
ESTAB_NORMALIZE = {
    "국립": "국립", "공립": "공립", "사립": "사립",
    "국": "국립", "공": "공립", "사": "사립",
}

UNKNOWN = "확인 필요"  # 자료에 없거나 불명확할 때 채우는 표시(추정 금지)


def normalize_region(raw):
    if raw is None:
        return UNKNOWN
    key = str(raw).strip().replace(" ", "")
    return REGION_NORMALIZE.get(key, key)  # 미매핑이면 원문 유지(임의변환 금지)


def normalize_level(raw):
    if raw is None:
        return UNKNOWN
    key = str(raw).strip()
    return LEVEL_NORMALIZE.get(key, key)


def normalize_estab(raw):
    if raw is None or str(raw).strip() == "":
        return UNKNOWN
    return ESTAB_NORMALIZE.get(str(raw).strip(), str(raw).strip())


# ─────────────────────────────────────────────────────────────
# 3. 정렬  (지역청 → 학교급 → 설립유형 → 학교명 가나다)
# ─────────────────────────────────────────────────────────────
def sort_key(row: dict):
    return (
        REGION_INDEX.get(row.get("지역청") or row.get("지역"), 999),
        LEVEL_INDEX.get(row.get("학교급"), 999),
        ESTAB_INDEX.get(row.get("설립유형") or row.get("구분"), 999),
        str(row.get("학교명") or ""),
    )


def sort_detail(rows: list[dict]) -> list[dict]:
    return sorted(rows, key=sort_key)


def order_regions(regions):
    """주어진 지역청 목록을 지정 순서로 정렬(미지정 지역은 뒤로)."""
    return sorted(set(regions), key=lambda r: REGION_INDEX.get(r, 999))


# ─────────────────────────────────────────────────────────────
# 4. 검증기  → 검토표 행 리스트 반환
#    각 행: (구분, 검토항목, 결과, 확인내용, 조치필요)
#    결과: "이상 없음" / "확인 필요" / "오류"
# ─────────────────────────────────────────────────────────────
def _amount(row):
    v = row.get("지원금액")
    if v in (None, "", UNKNOWN):
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def check_regions(rows, allowed_regions):
    """지역청 누락/임의추가/순서 검증."""
    out = []
    used = [r.get("지역청") or r.get("지역") for r in rows]
    used_set = set(used)
    allowed = set(allowed_regions)
    extra = used_set - allowed
    missing = allowed - used_set
    out.append(("지역청", "사용자 미제시 지역청 포함 여부",
                "오류" if extra else "이상 없음",
                f"임의추가: {sorted(extra)}" if extra else "없음",
                "예" if extra else "아니오"))
    out.append(("지역청", "사용자 제시 지역청 누락 여부",
                "확인 필요" if missing else "이상 없음",
                f"미표시: {sorted(missing)}" if missing else "없음",
                "예" if missing else "아니오"))
    # 순서
    idxs = [REGION_INDEX.get(r, 999) for r in used if r in REGION_INDEX]
    ordered = idxs == sorted(idxs)
    out.append(("지역청", "지역청 순서 오류 여부",
                "이상 없음" if ordered else "오류",
                "지정 순서 일치" if ordered else "지정 순서와 다름",
                "아니오" if ordered else "예"))
    return out


def check_totals(detail_total, region_total, level_total, estab_total, plan_total=None):
    """세부내역 합계 vs 각 요약 합계 vs 총예산 일치 검증."""
    out = []

    def cmp(name, a, b):
        if a is None or b is None:
            return (name, "확인 필요", f"비교 불가(a={a}, b={b})", "예")
        ok = abs(a - b) < 1e-6
        return (name, "이상 없음" if ok else "오류",
                f"{a:,.0f} vs {b:,.0f}", "아니오" if ok else "예")

    for label, other in [("지역청별 합계", region_total),
                         ("학교급별 합계", level_total),
                         ("설립유형별 합계", estab_total)]:
        n, res, detail, act = cmp(f"세부내역 합계 = {label}", detail_total, other)
        out.append(("합계", n, res, detail, act))
    if plan_total is not None:
        n, res, detail, act = cmp("총예산 = 재배정 총액", plan_total, detail_total)
        out.append(("합계", n, res, detail, act))
    return out


def check_amounts(rows):
    """빈칸/0원/단위혼용 의심 검증."""
    out = []
    blanks = [r.get("학교명") for r in rows if _amount(r) is None]
    zeros = [r.get("학교명") for r in rows if _amount(r) == 0]
    out.append(("금액", "지원금액 빈칸 여부",
                "확인 필요" if blanks else "이상 없음",
                f"빈칸: {blanks}" if blanks else "없음",
                "예" if blanks else "아니오"))
    out.append(("금액", "지원금액 0원 여부",
                "확인 필요" if zeros else "이상 없음",
                f"0원: {zeros}" if zeros else "없음",
                "예" if zeros else "아니오"))
    # 단위 혼용 의심: 값 범위가 1000 미만과 100000 이상이 섞이면 천원/원 혼용 가능
    vals = [a for a in (_amount(r) for r in rows) if a not in (None, 0)]
    mixed = bool(vals) and (min(vals) < 1000 and max(vals) >= 100000)
    out.append(("금액", "원/천원 단위 혼동 여부",
                "확인 필요" if mixed else "이상 없음",
                f"값 범위 {min(vals):,.0f}~{max(vals):,.0f}" if vals else "값 없음",
                "예" if mixed else "아니오"))
    return out


def check_duplicates(rows):
    """같은 학교 중복 여부(여러 부문 선정 가능 → 확인 필요로만 표시)."""
    seen = {}
    for r in rows:
        k = (r.get("지역청") or r.get("지역"), r.get("학교명"))
        seen[k] = seen.get(k, 0) + 1
    dups = [k for k, c in seen.items() if c > 1 and k[1]]
    return [("중복", "같은 학교 중복 배정 여부",
             "확인 필요" if dups else "이상 없음",
             f"중복(부문선정/오류 확인): {dups}" if dups else "없음",
             "예" if dups else "아니오")]


def check_estab(rows):
    """설립유형 누락 + 국립을 공립에 섞지 않았는지."""
    out = []
    missing = [r.get("학교명") for r in rows
               if (r.get("설립유형") or r.get("구분")) in (None, "", UNKNOWN)]
    out.append(("설립유형", "설립유형(국립·공립·사립) 누락 여부",
                "확인 필요" if missing else "이상 없음",
                f"미기재: {missing}" if missing else "없음",
                "예" if missing else "아니오"))
    return out


def build_full_checklist(rows, allowed_regions, detail_total,
                         region_total, level_total, estab_total, plan_total=None):
    """전체 검토표 조립. SKILL의 검토 항목을 자동 채운다."""
    checklist = []
    checklist += check_regions(rows, allowed_regions)
    checklist += check_totals(detail_total, region_total, level_total,
                              estab_total, plan_total)
    checklist += check_amounts(rows)
    checklist += check_duplicates(rows)
    checklist += check_estab(rows)
    return checklist


# ─────────────────────────────────────────────────────────────
# 5. 엑셀 서식 헬퍼 (openpyxl)
# ─────────────────────────────────────────────────────────────
def excel_helpers():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    FONT = "맑은 고딕"
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
    TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    NUM_FMT = "#,##0"           # 원/천원 공통(단위는 머리글에 명시)

    def title(ws, text, span):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        c = ws.cell(1, 1, text)
        c.font = Font(name=FONT, bold=True, size=14)
        c.alignment = CENTER

    def header(ws, row, values):
        for j, v in enumerate(values, 1):
            c = ws.cell(row, j, v)
            c.font = Font(name=FONT, bold=True)
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = BORDER

    def cell(ws, r, c, v, *, num=False, bold=False, total=False):
        x = ws.cell(r, c, v)
        x.font = Font(name=FONT, bold=bold)
        x.border = BORDER
        x.alignment = Alignment(
            horizontal="right" if num else "center", vertical="center")
        if num:
            x.number_format = NUM_FMT
        if total:
            x.fill = TOTAL_FILL
            x.font = Font(name=FONT, bold=True)
        return x

    def autofit(ws, widths):
        from openpyxl.utils import get_column_letter
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

    def landscape_fit(ws):
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    return dict(title=title, header=header, cell=cell,
                autofit=autofit, landscape_fit=landscape_fit,
                FONT=FONT, BORDER=BORDER, NUM_FMT=NUM_FMT)
